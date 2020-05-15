# -*- coding: utf-8 -*-
#
from tkinter import *
from tkinter import font
import os
from PIL import Image, ImageTk
import openpyxl
import time
from pygame import mixer

def Achat_Pluie_Asteroides(score, inventaire):
    global worksheet, WPoints
    if score.value - 460 >= 0 and inventaire.value =='False':
        score.value = score.value - 460
        inventaire.value = 'True'
        Bravo = Wplan.create_text((largeur_fenetre // 2)+30 , 185, text="Bravo, vous venez d'acquérir la pluie d'astéroides !", fill='green', font=Times)
        Wplan.itemconfig(WPoints, text=str(B1.value)+" $")
        window.update()
        time.sleep(2)
        Wplan.delete(Bravo)
        worksheets.save(filename='Score.xlsx')
    elif inventaire.value == 'True' :
        Erreur=Wplan.create_text( (largeur_fenetre//2)+25, 185, text='Vous possédez déjà cet item', fill='red', font=Times)
        window.update()
        time.sleep(2)
        Wplan.delete(Erreur)
    else :
        Erreur = Wplan.create_text( (largeur_fenetre//2)+25, 185, text="Vous n'avez pas les moyens !", fill='red', font=Times)
        window.update()
        time.sleep(2)
        Wplan.delete(Erreur)


def ChangePage():
    os.startfile("Accueil.py")
    window.quit()
def Exit(evt):
    window.attributes('-fullscreen', True)
    Wplan.bind_all('<Escape>', Tixe)
def Tixe(evt):
    window.attributes('-fullscreen', False)
    Wplan.bind_all('<Escape>', Exit)

window=Tk()
window.title("SPACE JUMPER 3000")

#Faire en sorte que la taille de la fenêtre s'adapte à n'importe quel écran :
largeur_fenetre = window.winfo_screenwidth()
hauteur_fenetre = window.winfo_screenheight()

window.geometry("%dx%d+0+0" % (largeur_fenetre,hauteur_fenetre))
window.attributes('-fullscreen', True)

window.iconbitmap('Fusées/FuseeFeu3.ico')

#------- Police -------#

Times = font.Font(family='Times', size=24, weight='bold')
TimesTitre = font.Font(family='Times', size=36, weight='bold')
Times2 = font.Font(family='Times', size=16, weight='bold')

#Ciel=PhotoImage(file='CielAccueil.png')

#------Config Image-----#

Fusee =PhotoImage(file='Fusées/Fusee0.png')
FuseeFeu = PhotoImage(file='Fusées/FuseeFeu0.png')
PluieAsteroide = PhotoImage(file='img_pluie/Pluie_Asteroides0.png')

Invisible = PhotoImage(file='invisible.png')
Titre = Image.open('Titre.png')
TkTitre= ImageTk.PhotoImage(Titre)

#------- Configuration du score --------#
worksheets = openpyxl.load_workbook("Score.xlsx")
worksheet = worksheets.active
B1 = worksheet.cell(row=1, column=2)
B2 = worksheet.cell(row=2, column=2)

#------------- Graphique accueil --------#

Wplan=Canvas(window, width=largeur_fenetre, height=hauteur_fenetre)
Wplan.grid()

WCCiel = Image.open('Ciel.png')
WCCiel = WCCiel.resize((largeur_fenetre, hauteur_fenetre), Image.ANTIALIAS)
Ciel = ImageTk.PhotoImage(WCCiel)
WCiel = Wplan.create_image((largeur_fenetre // 2) - 2, (hauteur_fenetre // 2) - 2, image=Ciel)

#------------- Fullscreen -----------#
Wplan.bind_all('<Escape>', Tixe)

#------------
WBoutique = Wplan.create_text((largeur_fenetre//2), 100, text = 'Boutique', font = TimesTitre, fill='White', activefill='Yellow')

WScore = Wplan.create_text(largeur_fenetre - 190,  35, text='Banque : ', font=Times, fill= 'yellow', activefill='white')
WPoints = Wplan.create_text( largeur_fenetre - 70, 35, text= str(B1.value)+" $", font = Times, fill='yellow', activefill='white')
Wplan.create_line(largeur_fenetre - 255, 57, largeur_fenetre - 20, 57, fill='white', activefill='yellow', width=4)

WCarre1= Wplan.create_rectangle(50, 250, 300, 450, outline ='yellow')
WPluie_Asteroides = Wplan.create_image(170, 350, image=PluieAsteroide)
Pluie_Asteroides = Wplan.create_text(175, 465, text = "Pluie d'astéroides : 450 $", font = Times2, fill='yellow', activefill='white')
Buy_Pluie_Asteroides = Button(window, width = 18,text='Acheter', font=Times2, command= lambda : Achat_Pluie_Asteroides(B1, B2))
Buy_Pluie_Asteroides.place(x=60, y =480)

WAccueil = Button(Wplan, text='Accueil', font=Times, command= lambda : ChangePage())
WAccueil.place(x= 70, y = 100)

window.mainloop()
