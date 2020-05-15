from tkinter import *
from tkinter import font
import random
from timeit import default_timer
import os
from PIL import Image, ImageTk
from pygame import mixer
import time
import openpyxl

mixer.init()
mixer.music.load("DuaLipa.mp3")
mixer.music.play(-1)
mixer.music.set_volume(0.2)

#-------------Fonctions-----------#

def ChangePage():
    global worksheets, B1
    B1.value = B1.value + NbrePoints
    worksheets.save(filename="Score.xlsx")
    os.startfile("Accueil.py")
    window.quit()

#---- Mouvement fusée ---#
def Haut(evt):
    global Wplan
    global Wfusee
    PosY = Wplan.coords(Wfusee)[1]  #PosY devient la variable contenant les coordonées en Y de la fusée
    if PosY -50 > 50: #On empêche la fusée de sortir du cadre
            # Si la position de la fusée en Y -50 ( on retire 50 car la fusée à une largeur de 100 ) est plus grande que 50 alors :
        Wplan.coords(Wfusee, Wplan.coords(Wfusee)[0], PosY -50)
    else:
        Wplan.coords(Wfusee, Wplan.coords(Wfusee)[0], PosY)
def Bas(evt):
    global Wplan
    global Wfusee
    PosY = Wplan.coords(Wfusee)[1]
    if PosY +83 < hauteur_fenetre:  # la fusée a une largeur de 175 pixel, d'où le + 83 (175//2)
                            #Si sa position en Y est inférieur à la hauteur de la fenêtre alors :
        Wplan.coords(Wfusee, Wplan.coords(Wfusee)[0], PosY + 65)
    else:
        Wplan.coords(Wfusee, Wplan.coords(Wfusee)[0], PosY) # Sinon la fusée ne descend pas plus bas
def Gauche(evt):
    global Wplan
    global Wfusee
    PosX = Wplan.coords(Wfusee)[0] #PosX devient la variable contenant les coordonées en X de la fusée
    if PosX - 83 > 0:
        Wplan.coords(Wfusee, PosX - 30, Wplan.coords(Wfusee)[1])
    else:
        Wplan.coords(Wfusee, PosX, Wplan.coords(Wfusee)[1])
def Droite(evt):
    global Wplan
    global Wfusee

    PosX=Wplan.coords(Wfusee)[0]
    if PosX + 30 < 400:  #Délimitation en X à 400 de manière à ce que la fusée ne se déplace pas là où les astéroides spawn

        Wplan.coords(Wfusee, PosX +30, Wplan.coords(Wfusee)[1])
    else:
        Wplan.coords(Wfusee, PosX, Wplan.coords(Wfusee)[1])

#-----Generer Asteroide------#
class Attaque_asteroide :
    def __init__(self, AsteroX,AsteroY, Wplan):

        self.Wplan = Wplan
        self.id = self.Wplan.create_image(AsteroX, AsteroY)#, image=mesImages[NumImage]) #grâce à cette class, chaque asteroide est identifiable avec self.id
        self.coup = 0

    def animation_2(self):
        global Wfusee
        global Barre
        global Wplan
        global PremiereCollision, DeuxiemeCollision
        global NumImage, mesImages, Wplan, EnsembleAsteroide, Niveau4
        global NbrePoints, B1, worksheets

        if Wplan.coords(self.id)[0] < 0 :
            Wplan.delete(self.id)
            EnsembleAsteroide.remove(self)
            if Niveau4:
                if PremiereCollision == False:
                    Wplan.itemconfig(Barre, image=DeuxCarre)
                    PremiereCollision = True
                elif DeuxiemeCollision == False:
                    Wplan.itemconfig(Barre, image=UnCarre)
                    DeuxiemeCollision = True
                elif DeuxiemeCollision == True:
                    B1.value = B1.value + NbrePoints
                    worksheets.save(filename="Score.xlsx")
                    os.startfile("GameOverPage.py")
                    window.quit()
        else :
            self.Wplan.move(self.id, -23, 0)
            NumImage += 1
            self.Wplan.itemconfig(self.id, image=mesImages[NumImage])
            if NumImage == 23:
                NumImage = 0
            if (Wplan.coords(Wfusee)[0]) + 72.5 >= (Wplan.coords(self.id)[0]) - 50 and (Wplan.coords(Wfusee)[0]) - 72.5 <= (Wplan.coords(self.id)[0]) + 50 :
                if (Wplan.coords(Wfusee)[1]) + 25 >= (Wplan.coords(self.id)[1]) - 50 > (Wplan.coords(Wfusee)[1]) - 25 or (Wplan.coords(Wfusee)[1]) - 25 <= (Wplan.coords(self.id)[1]) + 50 < (Wplan.coords(Wfusee)[1]) + 25 or(Wplan.coords(self.id)[1]) + 50 >= (Wplan.coords(Wfusee)[1]) + 25 > (Wplan.coords(Wfusee)[1]) - 25 >= (Wplan.coords(self.id)[1]) - 50:
                    if PremiereCollision == False:
                        Wplan.itemconfig(Barre, image=DeuxCarre)
                        PremiereCollision = True
                    elif DeuxiemeCollision == False:
                        Wplan.itemconfig(Barre, image=UnCarre)
                        DeuxiemeCollision = True
                    elif DeuxiemeCollision == True:
                        B1.value = B1.value + NbrePoints
                        worksheets.save(filename="Score.xlsx")
                        os.startfile("GameOverPage.py")
                        window.quit()
                    AsteroideDetruit = mixer.Sound("AsteroideDetruit.wav")
                    AsteroideDetruit.set_volume(0.5)
                    AsteroideDetruit.play()
                    Wplan.delete(self.id)
                    EnsembleAsteroide.remove(self)

def Create_asteroide():
    global Temps_Generer, NumImage, mesImages

    if Temps_Generer > 900 : #Si le temps de génération entre deux asteroides est supérieur à 900
        Temps_Generer-=100  #Alors le temps diminue de 100 ms
    AsteroX = random.randint(1000, largeur_fenetre-50) #Coordonées en X de l'asteroide générés aléatoirement entre 1000 et la largeur de la fenêtre -50 car la largeur de l'asteroide est de 100
    AsteroY = random.randint(50, hauteur_fenetre-50) #Pareillement pour les coordonées en Y
    Wasteroide = Attaque_asteroide(AsteroX, AsteroY, Wplan) #Creation de l'asteroide grâce à la class Attaque_asteroide
    EnsembleAsteroide.add(Wasteroide)   #On repertorie l'asteroide dans le set EnsembleAsteroide

    window.after(Temps_Generer, Create_asteroide) #Après le temps de generation entre deux asteroide on recréer un nouvel asteroide
def animation_asteroide():
    try :
        for Wasteroide in EnsembleAsteroide.copy(): #Pour les asteroides dans le set EnsembleAsteroide
            Wasteroide.animation_2() #on utilise la fonction animation_2 dans la class Attaque_asteroide pour les animer
        window.after(350, animation_asteroide)
    except :
        return()

#--------Generer Laser -------#
class Perso:
    def __init__(self, Cordox, Cordoy, Wplan, img):
        self.Wplan = Wplan
        self.coup = 0
        self.id = self.Wplan.create_image(Cordox, Cordoy, image=img)
    def recover(self):
        global Wplan
        self.id=self.Wplan.create_image(Cordox, Cordoy)
    def animation(self):
        global Wplan
        self.Wplan.move(self.id, 6, 0)

    def Colision(self):
        global NbrePoints, EnsembleAsteroide,EnsembleLaser, Wplan
        for Wasteroide in EnsembleAsteroide.copy():
            if Wplan.coords(self.id)[0] + 20 >= Wplan.coords(Wasteroide.id)[0] -50 and (Wplan.coords(self.id)[0]) - 20 <= (Wplan.coords(Wasteroide.id)[0]) + 50 :
                if (Wplan.coords(self.id)[1]) + 4 >= (Wplan.coords(Wasteroide.id)[1]) - 50 > (Wplan.coords(self.id)[1]) - 25 or (Wplan.coords(self.id)[1]) - 25 <= (Wplan.coords(Wasteroide.id)[1]) + 50 < (Wplan.coords(self.id)[1]) + 25 or (Wplan.coords(Wasteroide.id)[1]) + 50 >= (Wplan.coords(self.id)[1]) + 25 > (Wplan.coords(self.id)[1]) - 25 >= (Wplan.coords(Wasteroide.id)[1]) - 50:
                    Explosion = mixer.Sound("BruitageCraquement.wav")
                    Explosion.set_volume(0.5)
                    Explosion.play()
                    Wasteroide.coup=(Wasteroide.coup+1)
                    Wplan.delete(self.id)
                    EnsembleLaser.remove(self)
                    if Wasteroide.coup == 3:
                        AsteroideDetruit = mixer.Sound("AsteroideDetruit.wav")
                        AsteroideDetruit.set_volume(0.5)
                        AsteroideDetruit.play()
                        Wplan.delete(Wasteroide.id)
                        EnsembleAsteroide.remove(Wasteroide)
                        NbrePoints += 1
                        Wplan.itemconfig(CanvasPoints, text=('Nombre de Points : ', NbrePoints))
                    break
            elif Wplan.coords(self.id)[0]  >= largeur_fenetre:
                Wplan.delete(self.id)
                EnsembleLaser.remove(self)
                break

def Create_laser(event):
    global Wplan, Laser
    Cordo = Wplan.coords(Wfusee)
    LaserX = Cordo.pop(0) + 87.5
    LaserY = Cordo.pop(0)
    personnage = Perso(LaserX, LaserY, Wplan, Laser)
    EnsembleLaser.add(personnage)
    Bruitage = mixer.Sound("BruitageLaserWav.wav")
    Bruitage.set_volume(0.05)
    Bruitage.play()
def animation_laser():
    global EnsembleLaser
    for personnage in EnsembleLaser.copy():
        personnage.animation()
        personnage.Colision()
    window.after(9, animation_laser)

#------ Super Pouvoir : Pluie d'astéroides -------#
def PluieRedirect(evt):
    global Wplan, B2
    if B2.value =='True':
        Wplan.bind_all('k', EnleverSort)
        Wplan.bind_all('<Motion>', Attaque_Pluie_Asteroides)
    else:
        Erreur = Label(window, text="Vous n'avez pas de pouvoirs", fg='red', font=Times)
        Erreur.place(x=(largeur_fenetre // 2) - 180, y=175)
        Erreur.after(2000, Erreur.destroy)

def Attaque_Pluie_Asteroides(evt):
    global Delimitation, Wplan, Attaque, B2, worksheets, EnsembleAsteroide, Wpluie, Delimitation1
    Wplan.delete(Delimitation)
    Delimitation = Wplan.create_rectangle(evt.x-250, evt.y-250, evt.x+250, evt.y+250, outline='red', width = 3)
    Wplan.bind_all('<Button-1>',  Lanceur)

def Lanceur(evt):
    global Delimitation, Wplan, Attaque, B2, worksheets, Wpluie
    Wplan.unbind_all('<Button-1>')
    Wplan.unbind_all('<Motion>')
    B2.value = 'False'
    worksheets.save(filename="Score.xlsx")
    Wpluie = Wplan.create_image(Wplan.coords(Delimitation)[0] + 250, Wplan.coords(Delimitation)[1] + 250,image=mesImagesPluie[NumImagePluie])
    for i in range(4):
        Delimitation1.insert(i, Wplan.coords(Delimitation)[i])
    Wplan.delete(Delimitation)
    Animation_Pluie_Asteroides()

def Animation_Pluie_Asteroides():
    global Wplan, PluieAsteroide
    global mesImagesPluie, NumImagePluie, Wpluie
    NumImagePluie+=1
    Wplan.itemconfig(Wpluie, image=mesImagesPluie[NumImagePluie])
    if NumImagePluie == 5:
        Detruire_Sort_Pluie()
        Wplan.delete(Wpluie)
    else:
        Wplan.after(225, Animation_Pluie_Asteroides)

Delimitation1=[]
def Detruire_Sort_Pluie():
    global Wplan, EnsembleAsteroide, Delimitation, Delimitation1, NbrePoints
    for Wasteroide in EnsembleAsteroide.copy():
        if Delimitation1[0] <= Wplan.coords(Wasteroide.id)[0] <= Delimitation1[2] and Delimitation1[1] <= Wplan.coords(Wasteroide.id)[1] <= Delimitation1[3]:
            AsteroideDetruit = mixer.Sound("AsteroideDetruit.wav")
            AsteroideDetruit.set_volume(0.5)
            AsteroideDetruit.play()
            Wplan.delete(Wasteroide.id)
            EnsembleAsteroide.remove(Wasteroide)
            NbrePoints += 1
            Wplan.itemconfig(CanvasPoints, text=('Nombre de Points : ', NbrePoints))
    del Delimitation1

def Inventaire_sorts():
    global Wplan, Pluie_Asteroides_Icone, B2, WSorts, WSort_Pluie
    if B2.value =='True':
        Wplan.itemconfig(WSorts, text ='Sort(s) disponible(s) :', font=Times)
        Wplan.itemconfig(WSort_Pluie, image=Pluie_Asteroides_Icone)
    elif B2.value =='False':
        Wplan.delete(WSorts)
        Wplan.delete(WSort_Pluie)
    Wplan.after(500, Inventaire_sorts)

def EnleverSort(evt):
    global Wplan, Delimitation
    Wplan.unbind_all('<Button-1>')
    Wplan.unbind_all('<Motion>')
    Wplan.delete(Delimitation)
    Wplan.bind_all('k', PluieRedirect)

#------- Niveaux ----#
def Niveaux(temps):
    global Niveau1, Niveau2, Niveau3, Niveau4
    if temps != "00:01:38" and Niveau1==True:
        WNiveau1 = Wplan.create_text((largeur_fenetre // 2) - 180, 175, text='Niveau 1', fill='White', font=TimesTitre)
        Wplan.after(2000, NiveauPasser(WNiveau1))
        window.update()
        window.after(2000, NiveauPasser(WNiveau1))
        NiveauPasser(WNiveau1)
        Niveau1, Niveau2 = False, True

def NiveauPasser(level):
    global Wplan
    Wplan.delete(level)
#-------FullScreen-------#
def Exit(evt): #Même fonctions utilisées pour l'accueil
    window.attributes('-fullscreen', True)
    Wplan.bind_all('<Escape>', Exit_inverse)
def Exit_inverse(evt):
    window.attributes('-fullscreen', False)
    Wplan.bind_all('<Escape>', Exit)

#----

def chronometre():
    global Niveau1, Niveau2, Niveau3, Niveau4, Wplan, WNiveau1, WNiveau2, WNiveau3, WNiveau4
    now = default_timer() - start
    minutes, seconds = divmod(now, 60)
    hours, minutes = divmod(minutes, 60)
    str_time = "%d:%02d:%02d" % (hours, minutes, seconds)
    Wplan.itemconfigure(chrono, text=str_time, fill='white', font=Times)
    if Niveau1==True:
        Niveau1=False
        WNiveau1 = Wplan.create_text((largeur_fenetre // 2), 75, text='Niveau 1', fill='White', font=TimesTitre)
    elif hours==0 and minutes==0 and 3 < seconds < 4:
        Wplan.delete(WNiveau1)
    elif hours ==0 and minutes==1 and 38 < seconds <39:
        WNiveau2 = Wplan.create_text((largeur_fenetre // 2), 75, text='Niveau 2', fill='yellow', font=TimesTitre)
    elif hours ==0 and minutes ==1 and 41 < seconds < 42:
        Wplan.delete(WNiveau2)
    elif hours ==0 and minutes==2 and 4 < seconds <5:
        WNiveau3 = Wplan.create_text((largeur_fenetre // 2), 75, text='Niveau 3', fill='orange', font=TimesTitre)
    elif hours ==0 and minutes ==2 and 6 < seconds < 7:
        Wplan.delete(WNiveau3)
    elif hours ==0 and minutes== 3and 20 < seconds <21:
        WNiveau4 = Wplan.create_text((largeur_fenetre // 2), 75, text='Niveau 4', fill='red', font=TimesTitre)
        Niveau4 = True
    elif hours ==0 and minutes ==3 and 23 < seconds < 24:
        Wplan.delete(WNiveau4)

    window.after(1000, chronometre)

WNiveau1=0
WNiveau2=0
WNiveau3=0
WNiveau4=0
#---------------------------------#

window = Tk()
window.title("SPACE JUMPER 3000")

# Faire en sorte que la taille de la fenêtre s'adapte à n'importe quel écran :
largeur_fenetre = window.winfo_screenwidth()
hauteur_fenetre = window.winfo_screenheight()

window.geometry("%dx%d+0+0" % (largeur_fenetre, hauteur_fenetre))
window.attributes('-fullscreen', True)

#------- Police -------#
TimesTitre = font.Font(family='Times', size=32, weight='bold')
Times = font.Font(family='Times', size=18, weight='bold')
Times2 = font.Font(family='Times', size=16, weight='bold')

#------ Variable fonction create_laser et animation_laser -------#

EnsembleLaser=set() #Initialisation d'un set contenant les lasers créés par l'utilisateur

#------ Variables fonction create_asteroide et animation_asteroide -------#

EnsembleAsteroide=set() #Initialisation d'un set contenant les astéroides créés par le programme
Temps_Generer=5000 # Temps de génération entre les deux premiers asteroides de  5 secondes (diminue dans le temps)

# --------- #
animation_laser()
animation_asteroide()

#---------Image-------#
Fusee=PhotoImage(file='Fusées/Fusee0.png')
Laser=PhotoImage(file='Laser.png')
Invisible=PhotoImage(file='invisible.png')
Pluie_Asteroides_Icone=PhotoImage(file="img_pluie/Pluie_Asteroides_Icone.png")

window.iconbitmap('Fusées/FuseeFeu3.ico')

#------------- Graphique accueil --------#
Wplan=Canvas(window, width=largeur_fenetre, height=hauteur_fenetre)
Wplan.grid()
#------ Fond --------- #
WCCiel=Image.open('Ciel.png')
WCCiel = WCCiel.resize((largeur_fenetre, hauteur_fenetre), Image.ANTIALIAS) # La taille du fond s'adapte à la taille de l'ecran grâce à la bibliothèque PIL
Ciel=ImageTk.PhotoImage(WCCiel)   #Configuration de l'image pour tkinter
WCiel=Wplan.create_image((largeur_fenetre//2)-2, (hauteur_fenetre//2)-2, image=Ciel) #Creation de l'image dans le canvas

#-------#
AffichageChrono = Wplan.create_text(largeur_fenetre//2,hauteur_fenetre//2)

#------#
Wfusee=Wplan.create_image(100,360, image=Fusee) # Affichage de la fusée

#---- Variables pour animation des astéroides ---#
mesImages=[] # Liste qui contient les positions de l'asteroide et servira a animer l'asteroide
for i in range(24):
    nom = "img\Astéroïde" + str(i) + ".png"
    mesImages.append(PhotoImage(file=nom))
NumImage=0

mesImagesPluie=[]
for i in range(6):
    nom ="img_Pluie/Pluie_Asteroides"+ str(i)+".png"
    mesImagesPluie.append(PhotoImage(file=nom))
NumImagePluie=0
Wpluie=0
WSort_Pluie=Wplan.create_image(325, 45)
WSorts=Wplan.create_text(190, 40, fill='yellow', activefill ='white')

#---Affichage Chrono ---#
start = default_timer()
chrono = Wplan.create_text(40, 20)

#--Nombres de Points---#
NbrePoints = 0
CanvasPoints = Wplan.create_text(largeur_fenetre/2, 20, text=('Nombre de Points :', NbrePoints), fill='white', font=Times)

#-----Déplacement fusée -----#
Wplan.bind_all('z', Haut) # Lorsque l'utilisateur appuie sur la touche z (minuscule) la fonction Haut s'active, la fusée monte
Wplan.bind_all('s', Bas)
Wplan.bind_all('d', Droite)
Wplan.bind_all('q', Gauche)

#------Plein écran ----------#
Wplan.bind_all('<Escape>', Exit_inverse)

#----Tir laser--------#
window.bind('<space>', Create_laser)

#---- Super pouvoirs ----#
Wplan.bind_all('k', PluieRedirect)
Delimitation = 0

worksheets = openpyxl.load_workbook("Score.xlsx")
worksheet = worksheets.active
B1 = worksheet.cell(row=1, column=2)
B2 = worksheet.cell(row=2, column=2) # Verification de la posséssion du pouvoir

#---Barre de Vie---#
TroisCarre=PhotoImage(file='img/Barre de Progression/3.png')
Barre = Wplan.create_image(largeur_fenetre*0.957, hauteur_fenetre*0.026, image=TroisCarre)
DeuxCarre=PhotoImage(file='img/Barre de Progression/2.png')
UnCarre=PhotoImage(file='img/Barre de Progression/1.png')
GameOver=PhotoImage(file='img/Barre de Progression/Game Over.png')

#--- Variables pour collision, fonction animation_asteroide
PremiereCollision = False #--- Variables booleennes qui serviront à faire baisser la barre de vie du joueur lors d'une collision avec un asteroide
DeuxiemeCollision = False

#----- Niveaux ----#:
Niveau1 = True
Niveau2 =False
Niveau3 = False
Niveau4= False

#----- Quitter ----#
WQuitter = Button(Wplan, text='Accueil', command= lambda : ChangePage(), font=Times)
WQuitter.place(x=20, y = hauteur_fenetre - 65)

#----
chronometre()
Create_asteroide()
Inventaire_sorts()

window.mainloop()